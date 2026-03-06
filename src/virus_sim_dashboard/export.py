"""Module for exporting the final simulation configuration to an Excel file."""

from base64 import b64decode
from io import BytesIO

import openpyxl as oxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows as df_to_rows
from openpyxl.worksheet.worksheet import Worksheet


def merge_cells(ws: Worksheet, headers: dict, col_name: str, row_start: int, n_rows: int):
    """Merge cells in the specified column for the given number of rows.

    Args:
        ws: The worksheet object.
        headers: A dictionary mapping column names to their respective column numbers.
        col_name: The name of the column to merge.
        row_start: The starting row number for the merge.
        n_rows: The number of rows to merge.
    """
    col_num = headers[col_name]
    if n_rows > 1:
        ws.merge_cells(
            start_row=row_start,
            start_column=col_num,
            end_row=row_start + n_rows - 1,
            end_column=col_num,
        )


def export_xlsx(config, filename="model_export.xlsx") -> BytesIO:
    """Export the simulation configuration to an Excel file.

    Args:
        config: The simulation configuration dictionary containing the fitted LoS distributions.
        filename: The name of the Excel file to save the exported data to.
    """
    # Note that Steps 1 and 2 are only used for determining patient start times when fitting
    # LoS distributions in Step 3, so they are omitted from the export.

    # Record fitted LoS distributions from Step 3: GIM patients

    wb = oxl.Workbook()

    # For GIM patients, use the existing first sheet in the workbook
    gim_data = config["step3"]["los_fit_results"]["gim"]
    ws = wb.active
    ws.title = "GIM_patients"
    write_xlsx_gim(ws, gim_data)

    # For ICU patients, create a new sheet in the workbook
    icu_data = config["step3"]["los_fit_results"]["icu"]
    ws = wb.create_sheet(title="ICU_patients")
    write_xlsx_icu(ws, icu_data)

    # Save jitter value (and other relevant parameters if any) from the config to a separate sheet
    params_ws = wb.create_sheet(title="Parameters")
    params_ws["A1"] = "Jitter"
    params_ws["B1"] = config["step4"]["jitter"] / 100  # Convert to percentage
    params_ws["B1"].number_format = "0%"

    # Record the daily patient counts from Step 4
    ws = wb.create_sheet(title="Dailies")
    dailies_df = pd.read_feather(BytesIO(b64decode(config["step4"]["dailies"])))
    for r in df_to_rows(dailies_df, index=False, header=True):
        ws.append(r)

    # Bold the header row (only the filled part)
    n_cols = len(dailies_df.columns)
    for col_num in range(1, n_cols + 1):
        ws.cell(row=1, column=col_num).font = oxl.styles.Font(bold=True)

    # Record the hourly patient distributions from Step 4
    ws = wb.create_sheet(title="Hourlies")
    hourlies_df = pd.read_feather(BytesIO(b64decode(config["step4"]["hourlies"])))
    for r in df_to_rows(hourlies_df, index=False, header=True):
        ws.append(r)

    # Set the number format for the probability columns to percentage with 3 decimal places
    # i.e B2:B25 (24 hours)
    for row_num in range(2, 26):
        ws.cell(row=row_num, column=2).number_format = "0.000%"

    # Bold the header row (only the filled part)
    n_cols = len(hourlies_df.columns)
    for col_num in range(1, n_cols + 1):
        ws.cell(row=1, column=col_num).font = oxl.styles.Font(bold=True)

    bytes_io = BytesIO()
    wb.save(bytes_io)
    bytes_io.seek(0)
    return bytes_io


def write_xlsx_gim(ws: Worksheet, gim_data: dict):
    """Write the GIM patient data to the given worksheet starting from the specified row number.

    Args:
        ws: The worksheet object to write to.
        gim_data: The dictionary containing the GIM patient data to write.
    """
    # List of all column headers for GIM patients
    _gim_headers = [
        "Pathway",
        "Label",
        "# Patients",
        "Probability",
        "Outcome",
        "Age",
        "Count",
        "% total",
        "% in label",
        "Distribution",
        "Param1",
        "Param2",
        "Param3",
    ]
    gim_headers = {x[1]: x[0] for x in list(enumerate(_gim_headers, start=1))}

    # Headers to merge for GIM patients, per label group
    gim_headers_merge = [
        "Pathway",
        "Label",
        "# Patients",
        "Probability",
        "Distribution",
        "Param1",
        "Param2",
        "Param3",
    ]

    # Write the column headers
    for header, col_num in gim_headers.items():
        ws.cell(row=1, column=col_num, value=header)
        ws.cell(row=1, column=col_num).font = oxl.styles.Font(bold=True)

    # Write the data rows starting from Row 2
    row_num = 2

    for label, label_data in gim_data.items():
        n_groups = len(label_data["label_groups"])

        if n_groups > 1:
            for col_name in gim_headers_merge:
                merge_cells(ws, gim_headers, col_name=col_name, row_start=row_num, n_rows=n_groups)

        # Write the label-level data to the first row for this label
        ws.cell(row=row_num, column=gim_headers["Pathway"], value="GIM")
        ws.cell(row=row_num, column=gim_headers["Label"], value=label)
        ws.cell(row=row_num, column=gim_headers["# Patients"], value=label_data["n_patients"])

        ws.cell(row=row_num, column=gim_headers["Probability"], value=label_data["probability"])
        ws.cell(row=row_num, column=gim_headers["Probability"]).number_format = "0.000%"

        ws.cell(
            row=row_num,
            column=gim_headers["Distribution"],
            value=label_data["los_fit"]["distribution"],
        )
        ws.cell(
            row=row_num, column=gim_headers["Param1"], value=label_data["los_fit"]["parameters"][0]
        )
        ws.cell(
            row=row_num, column=gim_headers["Param2"], value=label_data["los_fit"]["parameters"][1]
        )
        ws.cell(
            row=row_num, column=gim_headers["Param3"], value=label_data["los_fit"]["parameters"][2]
        )
        ws.cell(row=row_num, column=gim_headers["Param1"]).number_format = "0.0000"
        ws.cell(row=row_num, column=gim_headers["Param2"]).number_format = "0.0000"
        ws.cell(row=row_num, column=gim_headers["Param3"]).number_format = "0.0000"

        # Write the group-level data for each group in the label
        for group, group_data in label_data["label_groups"].items():
            # key is in the format "("gim", '{outcome}', '{age_group}')"
            _, outcome, age_group = [s.strip("'") for s in group.strip("()").split(", ")]
            ws.cell(row=row_num, column=gim_headers["Outcome"], value=outcome)
            ws.cell(row=row_num, column=gim_headers["Age"], value=age_group)
            ws.cell(row=row_num, column=gim_headers["Count"], value=group_data["count"])
            ws.cell(row=row_num, column=gim_headers["% total"], value=group_data["p_total"])
            ws.cell(row=row_num, column=gim_headers["% total"]).number_format = "0.000%"
            ws.cell(row=row_num, column=gim_headers["% in label"], value=group_data["p_in_label"])
            ws.cell(row=row_num, column=gim_headers["% in label"]).number_format = "0.000%"
            row_num += 1


def write_xlsx_icu(ws: Worksheet, icu_data: dict):
    """Write the ICU patient data to the given worksheet starting from the specified row number.

    Args:
        ws: The worksheet object to write to.
        icu_data: The dictionary containing the ICU patient data to write.
    """
    # Write the column headers
    _icu_headers = [
        "Pathway",
        "Label",
        "# Patients",
        "Probability",
        "Outcome",
        "Age",
        "Count",
        "% total",
        "% in label",
        "Pre_Distribution",
        "Pre_Param1",
        "Pre_Param2",
        "Pre_Param3",
        "Distribution",
        "Param1",
        "Param2",
        "Param3",
        "Post_Distribution",
        "Post_Param1",
        "Post_Param2",
        "Post_Param3",
    ]
    icu_headers = {x[1]: x[0] for x in list(enumerate(_icu_headers, start=1))}

    icu_headers_no_merge = ["Outcome", "Age", "Count", "% total", "% in label"]

    # Write the headers to the first row
    for header, col_num in icu_headers.items():
        ws.cell(row=1, column=col_num, value=header)
        ws.cell(row=1, column=col_num).font = oxl.styles.Font(bold=True)

    # Write the data rows starting from Row 2
    row_num = 2

    for label, label_data in icu_data.items():
        n_groups = len(label_data["label_groups"])

        if n_groups > 1:
            for col_name in _icu_headers:
                if col_name not in icu_headers_no_merge:
                    merge_cells(
                        ws, icu_headers, col_name=col_name, row_start=row_num, n_rows=n_groups
                    )

        # Write the label-level data to the first row for this label
        ws.cell(row=row_num, column=icu_headers["Pathway"], value="ICU")
        ws.cell(row=row_num, column=icu_headers["Label"], value=label)
        ws.cell(row=row_num, column=icu_headers["# Patients"], value=label_data["n_patients"])

        ws.cell(row=row_num, column=icu_headers["Probability"], value=label_data["probability"])
        ws.cell(row=row_num, column=icu_headers["Probability"]).number_format = "0.000%"

        ws.cell(
            row=row_num,
            column=icu_headers["Pre_Distribution"],
            value=label_data["pre_icu_los_fit"]["distribution"],
        )
        ws.cell(
            row=row_num,
            column=icu_headers["Pre_Param1"],
            value=label_data["pre_icu_los_fit"]["parameters"][0],
        )
        ws.cell(
            row=row_num,
            column=icu_headers["Pre_Param2"],
            value=label_data["pre_icu_los_fit"]["parameters"][1],
        )
        ws.cell(
            row=row_num,
            column=icu_headers["Pre_Param3"],
            value=label_data["pre_icu_los_fit"]["parameters"][2],
        )
        ws.cell(row=row_num, column=icu_headers["Pre_Param1"]).number_format = "0.0000"
        ws.cell(row=row_num, column=icu_headers["Pre_Param2"]).number_format = "0.0000"
        ws.cell(row=row_num, column=icu_headers["Pre_Param3"]).number_format = "0.0000"

        ws.cell(
            row=row_num,
            column=icu_headers["Distribution"],
            value=label_data["icu_los_fit"]["distribution"],
        )
        ws.cell(
            row=row_num,
            column=icu_headers["Param1"],
            value=label_data["icu_los_fit"]["parameters"][0],
        )
        ws.cell(
            row=row_num,
            column=icu_headers["Param2"],
            value=label_data["icu_los_fit"]["parameters"][1],
        )
        ws.cell(
            row=row_num,
            column=icu_headers["Param3"],
            value=label_data["icu_los_fit"]["parameters"][2],
        )
        ws.cell(row=row_num, column=icu_headers["Param1"]).number_format = "0.0000"
        ws.cell(row=row_num, column=icu_headers["Param2"]).number_format = "0.0000"
        ws.cell(row=row_num, column=icu_headers["Param3"]).number_format = "0.0000"

        ws.cell(
            row=row_num,
            column=icu_headers["Post_Distribution"],
            value=label_data["post_icu_los_fit"]["distribution"],
        )
        ws.cell(
            row=row_num,
            column=icu_headers["Post_Param1"],
            value=label_data["post_icu_los_fit"]["parameters"][0],
        )
        ws.cell(
            row=row_num,
            column=icu_headers["Post_Param2"],
            value=label_data["post_icu_los_fit"]["parameters"][1],
        )
        ws.cell(
            row=row_num,
            column=icu_headers["Post_Param3"],
            value=label_data["post_icu_los_fit"]["parameters"][2],
        )
        ws.cell(row=row_num, column=icu_headers["Post_Param1"]).number_format = "0.0000"
        ws.cell(row=row_num, column=icu_headers["Post_Param2"]).number_format = "0.0000"
        ws.cell(row=row_num, column=icu_headers["Post_Param3"]).number_format = "0.0000"

        # Write the group-level data for each group in the label
        for group, group_data in label_data["label_groups"].items():
            # key is in the format "("icu", '{outcome}', '{age_group}')"
            _, outcome, age_group = [s.strip("'") for s in group.strip("()").split(", ")]
            ws.cell(row=row_num, column=icu_headers["Outcome"], value=outcome)
            ws.cell(row=row_num, column=icu_headers["Age"], value=age_group)
            ws.cell(row=row_num, column=icu_headers["Count"], value=group_data["count"])
            ws.cell(row=row_num, column=icu_headers["% total"], value=group_data["p_total"])
            ws.cell(row=row_num, column=icu_headers["% total"]).number_format = "0.000%"
            ws.cell(row=row_num, column=icu_headers["% in label"], value=group_data["p_in_label"])
            ws.cell(row=row_num, column=icu_headers["% in label"]).number_format = "0.000%"
            row_num += 1
