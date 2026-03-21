"""Functions to import data from Excel files."""

import typing

from openpyxl.worksheet.worksheet import Worksheet


def read_xlsx_gim(ws: Worksheet) -> dict[str, typing.Any]:
    """Parse the GIM patient parameters from the given worksheet.

    TODO: validate probabilities sum to 1 for each relevant population, or recalculate them.
    """
    range_sizes = {r.start_cell: r.size["rows"] for r in ws.merged_cells.ranges}
    curr_row = 2  # Skip header row
    result = {}
    while True:
        # A{row} is always "GIM"  -- if it's empty, we've reached the end of the table
        if ws[f"A{curr_row}"].value is None:
            break

        label = str(ws[f"B{curr_row}"].value)
        n_patients_label = int(ws[f"C{curr_row}"].value)
        prob_label = float(ws[f"D{curr_row}"].value)
        # columns E to I are per group, we will return to them later
        dist_name = str(ws[f"J{curr_row}"].value)
        dist_params = [float(ws[f"{col}{curr_row}"].value) for col in ["K", "L", "M"]]

        # How many groups are there for this label? Count the number of merged cells in column A
        n_groups = 1
        if ws[f"A{curr_row}"] in range_sizes:
            n_groups = range_sizes[ws[f"A{curr_row}"]]

        # Read the group-specific parameters from columns E to I
        group_params = {}
        for group_idx in range(n_groups):
            key = (
                "gim",
                str(ws[f"E{curr_row + group_idx}"].value),  # outcome
                str(ws[f"F{curr_row + group_idx}"].value),  # age group
            )
            group_params[f"{key}"] = {
                # Probability of this group in the full patient population (GIM + ICU)
                "p_total": float(ws[f"H{curr_row + group_idx}"].value),
                # Probability of this group among patients with the given label
                "p_in_label": float(ws[f"I{curr_row + group_idx}"].value),
            }

        # Advance the current row by the number of groups
        curr_row += n_groups

        result[label] = {
            "n_patients": n_patients_label,
            "probability": prob_label,
            "label_groups": group_params,
            "los_fit": {
                "distribution": dist_name,
                "parameters": dist_params,
            },
        }

    return result


def read_xlsx_icu(ws: Worksheet) -> dict[str, typing.Any]:
    """Parse the ICU patient parameters from the given worksheet.

    TODO: validate probabilities sum to 1 for each relevant population, or recalculate them.
    """
    range_sizes = {r.start_cell: r.size["rows"] for r in ws.merged_cells.ranges}
    curr_row = 2  # Skip header row
    result = {}
    while True:
        # A{row} is always "ICU"  -- if it's empty, we've reached the end of the table
        if ws[f"A{curr_row}"].value is None:
            break

        label = str(ws[f"B{curr_row}"].value)
        n_patients_label = int(ws[f"C{curr_row}"].value)
        prob_label = float(ws[f"D{curr_row}"].value)
        # columns E to I are per group, we will return to them later

        # Probability of GIM stay before/after ICU stay
        prob_pre_icu = float(ws[f"J{curr_row}"].value)
        prob_post_icu = float(ws[f"K{curr_row}"].value)

        # Pre-ICU LOS distribution in a GIM bed
        pre_dist_name = str(ws[f"L{curr_row}"].value)
        pre_dist_params = [float(ws[f"{col}{curr_row}"].value) for col in ["M", "N", "O"]]

        # ICU LOS distribution
        dist_name = str(ws[f"P{curr_row}"].value)
        dist_params = [float(ws[f"{col}{curr_row}"].value) for col in ["Q", "R", "S"]]

        # Post-ICU LOS distribution in a GIM bed
        post_dist_name = str(ws[f"T{curr_row}"].value)
        post_dist_params = [float(ws[f"{col}{curr_row}"].value) for col in ["U", "V", "W"]]

        # How many groups are there for this label? Count the number of merged cells in column A
        n_groups = 1
        if ws[f"A{curr_row}"] in range_sizes:
            n_groups = range_sizes[ws[f"A{curr_row}"]]

        # Read the group-specific parameters from columns E to I
        group_params = {}
        for group_idx in range(n_groups):
            key = (
                "icu",
                str(ws[f"E{curr_row + group_idx}"].value),  # outcome
                str(ws[f"F{curr_row + group_idx}"].value),  # age group
            )
            group_params[f"{key}"] = {
                # Probability of this group in the full patient population (GIM + ICU)
                "p_total": float(ws[f"H{curr_row + group_idx}"].value),
                # Probability of this group among patients with the given label
                "p_in_label": float(ws[f"I{curr_row + group_idx}"].value),
            }

        # Advance the current row by the number of groups
        curr_row += n_groups

        result[label] = {
            "n_patients": n_patients_label,
            "probability": prob_label,
            "label_groups": group_params,
            "prob_pre_icu": prob_pre_icu,
            "prob_post_icu": prob_post_icu,
            "pre_icu_los_fit": {
                "distribution": pre_dist_name,
                "parameters": pre_dist_params,
            },
            "icu_los_fit": {
                "distribution": dist_name,
                "parameters": dist_params,
            },
            "post_icu_los_fit": {
                "distribution": post_dist_name,
                "parameters": post_dist_params,
            },
        }

    return result
